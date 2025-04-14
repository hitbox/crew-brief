import io

from typing import Iterable

import openpyxl

from openpyxl.utils import get_column_letter

class WorkbookBuilder:

    def __init__(
        self,
        rowifier,
        hide_columns = None,
        freeze_panes = None,
        autofit_columns = False,
    ):
        self.rowifier = rowifier

        if hide_columns is None:
            hide_columns = []
        elif isinstance(hide_columns, Iterable):
            hide_columns = list(hide_columns)
        self.hide_columns = hide_columns

        self.freeze_panes = freeze_panes
        self.autofit_columns = autofit_columns

    def __call__(self, path_data, data):
        wb = openpyxl.Workbook()
        ws = wb.active

        for row in self.rowifier(data, path_data):
            ws.append(tuple(row))
            row.apply(ws)

        if self.freeze_panes:
            ws.freeze_panes = self.freeze_panes

        if self.hide_columns:
            for column in self.hide_columns:
                column = resolve_column(ws, column)
                ws.column_dimensions[column].hidden = True

        if self.autofit_columns:
            autofit_columns(ws)

        with io.BytesIO() as excel_stream:
            wb.save(excel_stream)
            excel_data = excel_stream.getvalue()

        return excel_data


def resolve_column(ws, col):
    """
    Resolves a column identifier using worksheet context.

    This is intended for values read from configuration or user input. Accepts
    either a numeric column index (1-based) or a string keyword like 'last' or
    'max_column', which will be resolved using worksheet data.
    """
    if isinstance(col, int):
        return col
    elif isinstance(col, str):
        col = col.lower()
        if col in ('last', 'max', 'max_column'):
            return get_column_letter(ws.max_column)
        raise ValueError(f'Unknown column keyword: {col}')
    else:
        raise ValueError(f'Invalid column identifier: {col}')

def autofit_columns(ws, padding=2):
    """
    Autofit columns.
    """
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            # Ignore merged cells.
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                # Ignore newline values for now.
                if isinstance(cell.value, str):
                    if '\n' in cell.value:
                        continue
                max_length = max(max_length, len(str(cell.value)))

        if col_letter is not None:
            # Resize with some padding
            ws.column_dimensions[col_letter].width = max_length + padding

