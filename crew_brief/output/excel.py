"""
Create Excel file from rowified UserEvents.txt (JSON) data.
"""

import datetime
import os

from itertools import repeat

import openpyxl

from openpyxl.utils import get_column_letter

try:
    import win32com.client as win32
except ImportError:
    win32 = None

from crew_brief import constants
from crew_brief.convert import to_excel_value

NB_SPACE = '\xa0'

muted_blue_fill = openpyxl.styles.PatternFill(
    # Muted blue
    start_color = 'dbeef4',
    end_color = 'dbeef4',
    fill_type = 'solid',
)

success_keys = set([
    'status',
])

def styles_from_workbook(source, prefix):
    wb = openpyxl.load_workbook(source)

class ExcelConverter:
    """
    Convert UserEvents JSON to Excel.
    """

    header = [
        'eventTimeStamp',
        'status',
        'eventType',
        'eventDetails',
        'Original eventDetails',
    ]
    first_three = header[:3]

    identifier_keys = set([
        'userId',
        'submittedBy',
    ])

    styles = {
        'header_style': {
            'font': openpyxl.styles.Font(
                bold = True,
                color = '000000',
            ),
            'alignment': openpyxl.styles.Alignment(
                vertical = 'center',
                horizontal = 'center',
                wrap_text = True,
            ),
            'fill': muted_blue_fill,
        },
        'field_value_header_style': {
            # Style side-by-side field-value pairs
            'font': openpyxl.styles.Font(
                bold = True,
                color = '000000',
            ),
            'alignment': openpyxl.styles.Alignment(
                vertical = 'center',
                horizontal = 'left',
            ),
            'fill': muted_blue_fill,
        },
        'key_style': {
            'font': openpyxl.styles.Font(
                bold = True,
            ),
            'fill': muted_blue_fill,
            #'alignment': openpyxl.styles.Alignment(
            #    vertical = 'top',
            #),
        },
        'value_style': {
            'font': openpyxl.styles.Font(
                #name = 'Courier New',
                name = 'Lucida Console',
            ),
        },
        'success_style': {
            'font': openpyxl.styles.Font(
                bold = True,
                color = '000000',
            ),
            'fill': openpyxl.styles.PatternFill(
                start_color = '7bf784',
                end_color = '7bf784',
                fill_type = 'solid',
            ),
            'alignment': openpyxl.styles.Alignment(
                horizontal = 'center',
            ),
        },
        'not_success_style': {
            'font': openpyxl.styles.Font(
                color = 'FFFFFF',
                bold = True,
            ),
            'fill': openpyxl.styles.PatternFill(
                start_color = 'FF0000',
                end_color = 'FF0000',
                fill_type = 'solid',
            ),
            'alignment': openpyxl.styles.Alignment(
                horizontal = 'center',
            ),
        },
        'list_style': {
            'alignment': openpyxl.styles.Alignment(
                wrap_text = True,
            ),
        },
        'event_details_header_style': {
            'alignment': openpyxl.styles.Alignment(
                horizontal = 'center',
            ),
        },
        'integer_style': {
            'number_format': '#,##0',
        },
        'float_style': {
            'number_format': '#,##0.00',
        },
        'full_datetime_style': {
            'number_format': 'yyyy-mm-dd hh:mm:ssZ',
        },
        'short_datetime_style': {
            'number_format': 'yyyy-mm-dd hh:mmZ',
        },
        'waypoint_style': {
            'font': openpyxl.styles.Font(
                bold = True,
            ),
            'fill': openpyxl.styles.PatternFill(
                start_color = 'FFFF00',
                end_color = 'FFFF00',
                fill_type = 'solid',
            ),
        },
        'identifier_style': {
            'alignment': openpyxl.styles.Alignment(
                horizontal = 'left',
            ),
        },
    }

    def __init__(self):
        self.style_names = [name for name in dir(self) if name.endswith('_style')]

    def autofit_columns(self, ws, padding=2):
        """
        Autofit columns.
        """
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                # Ignore merged cells.
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                if col_letter is None:
                    col_letter = cell.column_letter
                if cell.value:
                    # Ignore newline values for now.
                    if isinstance(cell.value, str):
                        if '\n' in cell.value:
                            continue
                    max_length = max(max_length, len(str(cell.value)))

            if col_letter is not None:
                # Resize with some padding
                ws.column_dimensions[col_letter].width = max_length + padding

    def autofit_rows(self, ws, font_height=12, line_spacing=1.05):
        """
        Autofit rows.

        :param ws: openpyxl Worksheet object.
        :param font_height: Font height in points.
        :param line_spacing: Extra padding in percent fraction.
        """

        def str_with_newlines(cell):
            return isinstance(cell.value, str) and '\n' in cell.value

        # Auto-fit row height based on newline count only if cells have
        # newlines.
        for row in ws.iter_rows():
            cells = [cell for cell in row if str_with_newlines(cell)]
            if not cells:
                continue
            max_lines = max(cell.value.count('\n') for cell in cells)
            if max_lines > 0:
                # Adjust factor as needed
                height = max_lines * font_height * line_spacing
                ws.row_dimensions[row[0].row].height = height

    def apply_style(self, cell, style):
        """
        Set attributes on a cell from data.
        """
        for attr, val in style.items():
            setattr(cell, attr, val)

    def get_styles_for_value(self, key, value, user_event_row):
        """
        Get styles to apply to a cell.
        """
        # Append styles to apply at end of loop.
        styles = []
        # Style for success field.
        if is_success(key, user_event_row):
            if value == 'success':
                styles.append('success_style')
            else:
                styles.append('not_success_style')
        # Style for types.
        if isinstance(value, (list, tuple)):
            styles.append('list_style')
        elif isinstance(value, int) and key not in self.identifier_keys:
            styles.append('integer_style')
        elif isinstance(value, float):
            styles.append('float_style')
        elif isinstance(value, datetime.datetime):
            styles.append('full_datetime_style')
        return styles

    def apply_row_styles(self, user_event_row, cells):
        """
        Apply styles to cells based on context from rowified data.
        """
        if user_event_row.keys:
            keys = user_event_row.keys
        else:
            keys = repeat(set())
        items = zip(keys, cells, user_event_row.row)
        for key, cell, value in items:
            styles = self.get_styles_for_value(key, value, user_event_row)
            # Apply styles.
            for style_name in styles:
                self.apply_style(cell, self.styles[style_name])

        # Style key-value pairs of rows.
        if 'field_and_value' in user_event_row.style_hint:
            self.apply_style(cells[0], self.styles['field_value_header_style'])

        if 'super-header' in user_event_row.style_hint:
            self.apply_style(cells[0], self.styles['header_style'])

        # Style the eventDetails sub-headers.
        if user_event_row.style_hint == 'user_event_fields_and_values header':
            for cell in cells[3:-1]:
                if cell.value is not None:
                    self.apply_style(cell, self.styles['header_style'])

    def append_header(self, ws, event_details_length):
        """
        Append header row and style.
        """
        # Build a header accounting for the width of eventDetails.
        header = self.header[:4]
        header.extend((None,)*(event_details_length-1))
        header.extend(self.header[4:])
        ws.append(header)

        # Merge eventDetails header. Check for zero because we just get the
        # calculated length of eventDetails which might be zero.
        if event_details_length - 1 > 0:
            ws.merge_cells(
                # Remember one-indexed column number.
                start_column = 4,
                end_column = 4 + event_details_length - 1,
                start_row = ws.max_row,
                end_row = ws.max_row,
            )

        # Apply header style to header row.
        for cell in ws[ws.max_row]:
            self.apply_style(cell, self.styles['header_style'])

    def __call__(
        self,
        rows,
        event_details_length,
        hide_event_details = False,
    ):
        """
        Convert UserEvents.txt JSON data to Excel.

        :param path_data: Data scraped from the path to the zip file.
        :param user_events_data: Original JSON data.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        # Freeze first row.
        ws.freeze_panes = 'A2'

        # Append header and styling.
        self.append_header(ws, event_details_length)

        # Append rows coming from rowifier and style. These can be headers and
        # values.
        for user_event_row in rows:
            row_for_excel = tuple(to_excel_value(value) for value in user_event_row.row)
            ws.append(row_for_excel)
            cells = ws[ws.max_row]
            self.apply_row_styles(user_event_row, cells)

            # Merge for the legIdentifier super-header.
            if user_event_row.style_hint == 'legIdentifier super-header':
                ws.merge_cells(
                    # Remember one-indexed column number.
                    start_column = 1,
                    end_column = 2,
                    start_row = ws.max_row,
                    end_row = ws.max_row,
                )

        # Autofit the columns. Autofitting the rows is very difficult and
        # spreadsheet applications seem to handle it if we leave it alone.
        self.autofit_columns(ws)

        if hide_event_details:
            last_column = get_column_letter(ws.max_column)
            ws.column_dimensions[last_column].hidden = True

        return wb


def autofit_with_excel(filename):
    """
    Autofit columns and rows with Excel.
    """
    excel = win32.Dispatch('Excel.Application')
    filename = os.path.abspath(os.path.normpath(filename))
    wb = excel.Workbooks.Open(filename)
    ws = wb.ActiveSheet
    ws.Cells.Columns.AutoFit()
    ws.Cells.Rows.AutoFit()
    wb.Save()

def is_success(key, user_event_row):
    """
    Test that a cell is the status field.
    """
    return (
        key in success_keys
        and
        user_event_row.style_hint in [
            'user_event_fields_and_values header',
            'user_event_fields_and_values singlerow',
        ]
    )
